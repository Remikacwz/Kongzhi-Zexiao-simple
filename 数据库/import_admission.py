# -*- coding: utf-8 -*-
'''
录取数据导入脚本
用法：
  1. 默认导入 SQLite（本地开发）：
       python import_admission.py
     会生成 admission.db、schools.csv、majors.csv、admissions.csv

  2. 导入 MySQL（线上服务器）：
       python import_admission.py --mysql \
         --host 127.0.0.1 --port 3306 --user root --password 123456 \
         --database kaoyan_admission
     需要先执行 schema_mysql.sql 建库建表。
'''
import argparse, csv, pathlib, re, sqlite3, sys, datetime
from db_config import mysql_config, sqlite_path

EXCEL_PATH = pathlib.Path(__file__).with_name('raw') / '27考研择校宝典_录取数据表_0815.xlsx'

HEADERS = [
    'school_name','college','major_text','planned_enrollment','retest_count','admitted_count',
    'retest_ratio','retest_max_score','retest_min_score','retest_avg_score','retest_politics_avg',
    'retest_english_subject','retest_english_avg','retest_math_subject','retest_math_avg',
    'retest_prof_subject','retest_prof_avg','admitted_max_score','admitted_min_score',
    'admitted_avg_score','admitted_politics_avg','admitted_english_subject','admitted_english_avg',
    'admitted_math_subject','admitted_math_avg','admitted_prof_subject','admitted_prof_avg',
]

def parse_major(text):
    text = (text or '').strip()
    m = re.match(r'^([0-9]{4}[0-9A-Za-z]{0,4})\s*(.*)$', text, re.S)
    if m:
        code, name = m.group(1), m.group(2).strip()
        return code, name, text
    return '', text, text

def to_num(v, cast):
    if v is None or str(v).strip() == '':
        return None
    try:
        return cast(str(v).strip())
    except Exception:
        return None

def compute_retest_ratio(retest_count, admitted_count, retest_ratio=None):
    """复录比为空时用 进复试人数/拟录取人数 补全；进复试人数小于拟录取人数时不补。"""
    if retest_ratio is None and retest_count is not None and admitted_count is not None and admitted_count > 0 and retest_count >= admitted_count:
        return round(retest_count / admitted_count + 1e-9, 2)
    return retest_ratio

def read_excel_rows(excel_path=EXCEL_PATH):
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def import_excel_to_db(excel_path, write_csv=True):
    '''import Excel to SQLite, return stats dict.'''
    excel_path = pathlib.Path(excel_path)
    conn = sqlite3.connect(sqlite_path())
    try:
        cur = conn.cursor()
        cur.executescript('''
        DROP TABLE IF EXISTS admissions;
        DROP TABLE IF EXISTS majors;
        DROP TABLE IF EXISTS schools;
        CREATE TABLE schools (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          name TEXT NOT NULL UNIQUE,
          province TEXT,
          tier TEXT,
          logo_url TEXT
        );
        CREATE TABLE majors (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          code TEXT NOT NULL,
          name TEXT NOT NULL,
          full_text TEXT NOT NULL,
          UNIQUE(code, name)
        );
        CREATE TABLE admissions (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          year INTEGER NOT NULL DEFAULT 2026,
          school_id INTEGER NOT NULL REFERENCES schools(id),
          major_id INTEGER NOT NULL REFERENCES majors(id),
          college TEXT,
          planned_enrollment INTEGER,
          retest_count INTEGER,
          admitted_count INTEGER,
          retest_ratio REAL,
          retest_max_score REAL,
          retest_min_score REAL,
          retest_avg_score REAL,
          retest_politics_avg REAL,
          retest_english_subject TEXT,
          retest_english_avg REAL,
          retest_math_subject TEXT,
          retest_math_avg REAL,
          retest_prof_subject TEXT,
          retest_prof_avg REAL,
          admitted_max_score REAL,
          admitted_min_score REAL,
          admitted_avg_score REAL,
          admitted_politics_avg REAL,
          admitted_english_subject TEXT,
          admitted_english_avg REAL,
          admitted_math_subject TEXT,
          admitted_math_avg REAL,
          admitted_prof_subject TEXT,
          admitted_prof_avg REAL,
          source_file TEXT,
          created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        ''')
        rows = read_excel_rows(excel_path)
        school_ids, major_ids, major_full = {}, {}, {}
        admission_rows = []
        for r in rows:
            d = dict(zip(HEADERS, r))
            school = (d.get('school_name') or '').strip()
            if not school:
                continue
            college = (d.get('college') or '').strip()
            major_text = (d.get('major_text') or '').strip()
            code, name, full_text = parse_major(major_text)
            if school not in school_ids:
                cur.execute('INSERT INTO schools(name) VALUES (?)', (school,))
                school_ids[school] = cur.lastrowid
            mkey = (code, name)
            if mkey not in major_ids:
                cur.execute('INSERT INTO majors(code, name, full_text) VALUES (?, ?, ?)', (code, name, full_text))
                major_ids[mkey] = cur.lastrowid
                major_full[mkey] = full_text
            retest_count = to_num(d.get('retest_count'), int)
            admitted_count = to_num(d.get('admitted_count'), int)
            retest_ratio = compute_retest_ratio(retest_count, admitted_count, to_num(d.get('retest_ratio'), float))
            vals = (
                2026, school_ids[school], major_ids[mkey], college,
                to_num(d.get('planned_enrollment'), int), retest_count,
                admitted_count, retest_ratio,
                to_num(d.get('retest_max_score'), float), to_num(d.get('retest_min_score'), float),
                to_num(d.get('retest_avg_score'), float), to_num(d.get('retest_politics_avg'), float),
                d.get('retest_english_subject'), to_num(d.get('retest_english_avg'), float),
                d.get('retest_math_subject'), to_num(d.get('retest_math_avg'), float),
                d.get('retest_prof_subject'), to_num(d.get('retest_prof_avg'), float),
                to_num(d.get('admitted_max_score'), float), to_num(d.get('admitted_min_score'), float),
                to_num(d.get('admitted_avg_score'), float), to_num(d.get('admitted_politics_avg'), float),
                d.get('admitted_english_subject'), to_num(d.get('admitted_english_avg'), float),
                d.get('admitted_math_subject'), to_num(d.get('admitted_math_avg'), float),
                d.get('admitted_prof_subject'), to_num(d.get('admitted_prof_avg'), float),
                excel_path.name,
            )
            cur.execute('''INSERT INTO admissions (
              year, school_id, major_id, college, planned_enrollment, retest_count, admitted_count,
              retest_ratio, retest_max_score, retest_min_score, retest_avg_score, retest_politics_avg,
              retest_english_subject, retest_english_avg, retest_math_subject, retest_math_avg,
              retest_prof_subject, retest_prof_avg, admitted_max_score, admitted_min_score,
              admitted_avg_score, admitted_politics_avg, admitted_english_subject, admitted_english_avg,
              admitted_math_subject, admitted_math_avg, admitted_prof_subject, admitted_prof_avg,
              source_file) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', vals)
            admission_rows.append((2026, school, code, name, college, *vals[4:]))
        conn.commit()
        if write_csv:
            school_rows = [(name,) for name in sorted(school_ids)]
            major_rows = [(k[0], k[1], major_full[k]) for k in sorted(major_ids)]
            with open(pathlib.Path(__file__).with_name('schools.csv'), 'w', newline='', encoding='utf-8-sig') as f:
                csv.writer(f).writerow(['school_name']); csv.writer(f).writerows(school_rows)
            with open(pathlib.Path(__file__).with_name('majors.csv'), 'w', newline='', encoding='utf-8-sig') as f:
                csv.writer(f).writerow(['major_code', 'major_name', 'full_text']); csv.writer(f).writerows(major_rows)
            with open(pathlib.Path(__file__).with_name('admissions.csv'), 'w', newline='', encoding='utf-8-sig') as f:
                csv.writer(f).writerow(['year', 'school_name', 'major_code', 'major_name', 'college'] + HEADERS[3:])
                csv.writer(f).writerows(admission_rows)
        return {'records': len(admission_rows), 'schools': len(school_ids), 'majors': len(major_ids)}
    finally:
        conn.close()


def import_sqlite():
    '''CLI：默认 Excel 导入 SQLite 并导出 CSV。'''
    info = import_excel_to_db(EXCEL_PATH, write_csv=True)
    print(f"SQLite OK: {info['records']} 条录取记录, {info['schools']} 所学校, {info['majors']} 个专业方向")
    print('已生成 admission.db / schools.csv / majors.csv / admissions.csv')


def import_mysql(host, port, user, password, database, excel_path=EXCEL_PATH):
    import pymysql
    conn = pymysql.connect(host=host, port=port, user=user, password=password, database=database, charset='utf8mb4')
    cur = conn.cursor()
    # 重建数据：先清空旧录取/专业/院校，再导入当前 Excel
    cur.execute('DELETE FROM admissions')
    cur.execute('DELETE FROM majors')
    cur.execute('DELETE FROM schools')
    rows = read_excel_rows(excel_path)
    school_ids, major_ids, major_full = {}, {}, {}
    n = 0
    for r in rows:
        d = dict(zip(HEADERS, r))
        school = (d.get('school_name') or '').strip()
        if not school:
            continue
        college = (d.get('college') or '').strip()
        code, name, full_text = parse_major((d.get('major_text') or '').strip())
        if school not in school_ids:
            cur.execute('INSERT INTO schools(name) VALUES (%s) ON DUPLICATE KEY UPDATE id=LAST_INSERT_ID(id)', (school,))
            school_ids[school] = cur.lastrowid
        mkey = (code, name)
        if mkey not in major_ids:
            cur.execute('INSERT INTO majors(code, name, full_text) VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE id=LAST_INSERT_ID(id)', (code, name, full_text))
            major_ids[mkey] = cur.lastrowid
        retest_count = to_num(d.get('retest_count'), int)
        admitted_count = to_num(d.get('admitted_count'), int)
        retest_ratio = compute_retest_ratio(retest_count, admitted_count, to_num(d.get('retest_ratio'), float))
        vals = (
            2026, school_ids[school], major_ids[mkey], college,
            to_num(d.get('planned_enrollment'), int), retest_count,
            admitted_count, retest_ratio,
            to_num(d.get('retest_max_score'), float), to_num(d.get('retest_min_score'), float),
            to_num(d.get('retest_avg_score'), float), to_num(d.get('retest_politics_avg'), float),
            d.get('retest_english_subject'), to_num(d.get('retest_english_avg'), float),
            d.get('retest_math_subject'), to_num(d.get('retest_math_avg'), float),
            d.get('retest_prof_subject'), to_num(d.get('retest_prof_avg'), float),
            to_num(d.get('admitted_max_score'), float), to_num(d.get('admitted_min_score'), float),
            to_num(d.get('admitted_avg_score'), float), to_num(d.get('admitted_politics_avg'), float),
            d.get('admitted_english_subject'), to_num(d.get('admitted_english_avg'), float),
            d.get('admitted_math_subject'), to_num(d.get('admitted_math_avg'), float),
            d.get('admitted_prof_subject'), to_num(d.get('admitted_prof_avg'), float),
            EXCEL_PATH.name,
        )
        cur.execute('''INSERT INTO admissions (
          year, school_id, major_id, college, planned_enrollment, retest_count, admitted_count,
          retest_ratio, retest_max_score, retest_min_score, retest_avg_score, retest_politics_avg,
          retest_english_subject, retest_english_avg, retest_math_subject, retest_math_avg,
          retest_prof_subject, retest_prof_avg, admitted_max_score, admitted_min_score,
          admitted_avg_score, admitted_politics_avg, admitted_english_subject, admitted_english_avg,
          admitted_math_subject, admitted_math_avg, admitted_prof_subject, admitted_prof_avg,
          source_file) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''', vals)
        n += 1
    conn.commit()
    cur.close(); conn.close()
    print(f'MySQL OK: {n} 条记录')


def import_mysql_from_config(excel_path=EXCEL_PATH):
    """根据 config.json 中的 mysql 配置导入 Excel 到 MySQL。"""
    mc = mysql_config()
    return import_mysql(
        host=mc.get('host', '127.0.0.1'),
        port=int(mc.get('port', 3306)),
        user=mc.get('user', 'root'),
        password=mc.get('password', ''),
        database=mc.get('database', 'kaoyan_admission'),
        excel_path=excel_path,
    )


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--mysql', action='store_true')
    ap.add_argument('--host', default='127.0.0.1')
    ap.add_argument('--port', type=int, default=3306)
    ap.add_argument('--user', default='root')
    ap.add_argument('--password', default='')
    ap.add_argument('--database', default='kaoyan_admission')
    args = ap.parse_args()
    if not EXCEL_PATH.exists():
        print('未找到 Excel:', EXCEL_PATH); sys.exit(1)
    if args.mysql:
        import_mysql(args.host, args.port, args.user, args.password, args.database)
    else:
        import_sqlite()
        try:
            import import_subjects
            subj = import_subjects.import_subjects_to_db(write_csv=True)
            print(f"subjects OK: {subj['subjects']} 门专业课, {subj['exam_subject_rows']} 条学校-科目, {subj['book_rows']} 条参考书目")
        except Exception as e:
            print('subject import skipped:', e)
